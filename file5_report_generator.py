"""
EnerCloud - File 5: Excel Report Generator
Pulls all historical data from Firebase and generates a professional
Excel report for government presentation.

SETUP:
    pip install firebase-admin openpyxl

USAGE:
    python file5_report_generator.py
    → Creates: EnerCloud_Energy_Report_<date>.xlsx
"""

import firebase_admin
from firebase_admin import credentials, db
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import datetime
import os

# ─── STEP 1: Replace with your Firebase Realtime Database URL ───────────────
DATABASE_URL = "https://enercloud-ed207-default-rtdb.firebaseio.com/"
# ────────────────────────────────────────────────────────────────────────────

CITIES = ["Mumbai", "Delhi", "Chennai", "Bangalore", "Hyderabad"]

# Color palette
DARK_BLUE  = "1F3864"
MID_BLUE   = "185FA5"
LIGHT_BLUE = "E6F1FB"
GREEN      = "1D9E75"
LIGHT_GREEN= "E1F5EE"
AMBER      = "BA7517"
LIGHT_AMBER= "FAEEDA"
RED        = "E24B4A"
GRAY_DARK  = "444441"
GRAY_LIGHT = "F1EFE8"
WHITE      = "FFFFFF"


def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def thin_border():
    side = Side(border_style="thin", color="D0D0D0")
    return Border(left=side, right=side, top=side, bottom=side)


def bold_font(size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=True, size=size, color=color, italic=italic)


def regular_font(size=10, color="000000"):
    return Font(name="Calibri", bold=False, size=size, color=color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ─── Fetch data from Firebase ─────────────────────────────────────────────────
def fetch_all_data():
    print("  Connecting to Firebase...")
    if not firebase_admin._apps:
        cred = credentials.Certificate("serviceAccountKey.json")
        firebase_admin.initialize_app(cred, {"databaseURL": DATABASE_URL})

    all_data = {}
    for city in CITIES:
        print(f"  Fetching data for {city}...")
        city_ref = db.reference(f"cities/{city}")
        history  = city_ref.child("history").get() or {}
        decisions= city_ref.child("decision_history").get() or {}
        latest   = city_ref.child("latest").get() or {}
        decision = city_ref.child("decision").get() or {}
        all_data[city] = {
            "history": list(history.values()) if history else [],
            "decisions": list(decisions.values()) if decisions else [],
            "latest": latest,
            "decision": decision,
        }
        count = len(all_data[city]["history"])
        print(f"    → {count} readings found")
    return all_data


# ─── Sheet 1: Cover Page ──────────────────────────────────────────────────────
def create_cover_sheet(wb, report_date):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 25

    # Header band
    for row in range(1, 8):
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].fill = hex_fill(DARK_BLUE)

    ws.merge_cells("A1:C7")
    ws["A1"].value = "EnerCloud Pvt. Ltd."
    ws["A1"].font = Font(name="Calibri", bold=True, size=28, color=WHITE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 100

    # Subtitle
    ws.merge_cells("A8:C8")
    ws["A8"].value = "Smart Grid Energy Management System"
    ws["A8"].font = bold_font(16, DARK_BLUE)
    ws["A8"].alignment = center()
    ws["A8"].fill = hex_fill(LIGHT_BLUE)
    ws.row_dimensions[8].height = 32

    # Metadata table
    meta = [
        ("Report Date:", report_date.strftime("%d %B %Y")),
        ("Report Type:", "Government Energy Infrastructure Report"),
        ("Prepared by:", "EnerCloud Analytics Engine v1.0"),
        ("Coverage:", "Mumbai, Delhi, Chennai, Bangalore, Hyderabad"),
        ("Data Source:", "Firebase Realtime Database (Google Cloud)"),
        ("Classification:", "Official Use Only"),
    ]
    for i, (label, value) in enumerate(meta):
        row = 10 + i * 2
        ws.row_dimensions[row].height = 22
        ws[f"B{row}"].value = label
        ws[f"B{row}"].font = bold_font(11, DARK_BLUE)
        ws[f"B{row}"].fill = hex_fill(GRAY_LIGHT)
        ws[f"B{row}"].alignment = left()
        ws[f"B{row}"].border = thin_border()

        ws[f"C{row}"].value = value
        ws[f"C{row}"].font = regular_font(11)
        ws[f"C{row}"].alignment = left()
        ws[f"C{row}"].border = thin_border()

    print("  ✓ Cover sheet created")


# ─── Sheet 2: City Summary ────────────────────────────────────────────────────
def create_summary_sheet(wb, all_data):
    ws = wb.create_sheet("City Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"].value = "Energy Summary — All Cities"
    ws["A1"].font = bold_font(14, WHITE)
    ws["A1"].fill = hex_fill(DARK_BLUE)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 36

    # Headers
    headers = ["City", "Readings\nCollected", "Avg Demand\n(kW)", "Avg Solar\n(kW)",
               "Solar Cover\n(%)", "Avg Battery\n(%)", "CO₂ Saved\n(kg)", "AI Decision"]
    widths   = [14, 12, 14, 14, 14, 14, 14, 22]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = bold_font(10, WHITE)
        cell.fill = hex_fill(MID_BLUE)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 36

    # Data rows
    for i, city in enumerate(CITIES):
        d = all_data[city]
        history = d["history"]
        decision = d["decision"]
        row = 3 + i

        if history:
            avg_demand  = sum(r.get("demand_kw",0)    for r in history) / len(history)
            avg_solar   = sum(r.get("solar_kw",0)     for r in history) / len(history)
            avg_batt    = sum(r.get("battery_soc_pct",0) for r in history) / len(history)
            solar_cover = (avg_solar / avg_demand * 100) if avg_demand > 0 else 0
        else:
            avg_demand = avg_solar = avg_batt = solar_cover = 0

        co2 = decision.get("co2_saved_kg", 0) if decision else 0
        ai_src = decision.get("recommended_source", "No data") if decision else "No data"

        row_data = [city, len(history), round(avg_demand,1), round(avg_solar,1),
                    round(solar_cover,1), round(avg_batt,1), co2, ai_src]

        fill_color = LIGHT_GREEN if i % 2 == 0 else WHITE
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = regular_font(10)
            cell.fill = hex_fill(fill_color)
            cell.alignment = center()
            cell.border = thin_border()

        ws.row_dimensions[row].height = 20

    # Bar chart: avg demand per city
    chart = BarChart()
    chart.type = "col"
    chart.title = "Average Energy Demand by City (kW)"
    chart.style = 10
    chart.y_axis.title = "kW"
    chart.x_axis.title = "City"

    data_ref = Reference(ws, min_col=3, min_row=2, max_row=2 + len(CITIES))
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=2 + len(CITIES))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    ws.add_chart(chart, "J2")

    print("  ✓ City Summary sheet created")


# ─── Sheet 3: Raw Data ────────────────────────────────────────────────────────
def create_raw_data_sheet(wb, all_data):
    ws = wb.create_sheet("Raw Sensor Data")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "Raw IoT Sensor Data — All Readings"
    ws["A1"].font = bold_font(12, WHITE)
    ws["A1"].fill = hex_fill(DARK_BLUE)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["City", "Timestamp", "Demand (kW)", "Solar (kW)",
               "Battery (%)", "Voltage (V)", "Frequency (Hz)", "Temp (°C)"]
    widths   = [14, 22, 12, 12, 12, 12, 14, 10]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = bold_font(10, WHITE)
        cell.fill = hex_fill(MID_BLUE)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    row = 3
    for city in CITIES:
        for reading in all_data[city]["history"]:
            fill_color = LIGHT_BLUE if row % 2 == 0 else WHITE
            row_vals = [
                city,
                reading.get("timestamp", "")[:19],
                reading.get("demand_kw", ""),
                reading.get("solar_kw", ""),
                reading.get("battery_soc_pct", ""),
                reading.get("grid_voltage_v", ""),
                reading.get("grid_frequency_hz", ""),
                reading.get("temperature_c", ""),
            ]
            for col, val in enumerate(row_vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = regular_font(9)
                cell.fill = hex_fill(fill_color)
                cell.alignment = center()
                cell.border = thin_border()
            ws.row_dimensions[row].height = 16
            row += 1

    print(f"  ✓ Raw Data sheet created ({row - 3} rows)")


# ─── Sheet 4: AI Decisions ────────────────────────────────────────────────────
def create_decisions_sheet(wb, all_data):
    ws = wb.create_sheet("AI Decisions")

    ws.merge_cells("A1:F1")
    ws["A1"].value = "AI Decision Engine — Recommendations Log"
    ws["A1"].font = bold_font(12, WHITE)
    ws["A1"].fill = hex_fill(DARK_BLUE)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    headers = ["City", "Recommended Source", "Solar Coverage (%)", "Efficiency (%)", "CO₂ Saved (kg)", "Alert"]
    widths   = [14, 22, 18, 14, 14, 40]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = bold_font(10, WHITE)
        cell.fill = hex_fill(DARK_BLUE)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    for i, city in enumerate(CITIES):
        d = all_data[city]["decision"] or {}
        row = 3 + i
        src = d.get("recommended_source", "No data")

        # Color-code by source
        if "Solar" in src and "+" not in src: fill = LIGHT_GREEN
        elif "Battery" in src: fill = "EEEDFE"
        elif "Grid" in src: fill = LIGHT_BLUE
        else: fill = LIGHT_AMBER

        row_vals = [city, src, d.get("solar_coverage_pct",""), d.get("system_efficiency_pct",""),
                    d.get("co2_saved_kg",""), d.get("alert","None")]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = regular_font(10)
            cell.fill = hex_fill(fill)
            cell.alignment = center()
            cell.border = thin_border()
        ws.row_dimensions[row].height = 20

    print("  ✓ AI Decisions sheet created")


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    report_date = datetime.datetime.now()
    filename = f"EnerCloud_Energy_Report_{report_date.strftime('%Y%m%d_%H%M')}.xlsx"

    print("=" * 60)
    print("  EnerCloud — Excel Report Generator")
    print("=" * 60)

    all_data = fetch_all_data()

    print("\n  Building Excel report...")
    wb = openpyxl.Workbook()

    create_cover_sheet(wb, report_date)
    create_summary_sheet(wb, all_data)
    create_raw_data_sheet(wb, all_data)
    create_decisions_sheet(wb, all_data)

    wb.save(filename)

    print(f"\n  ✅ Report saved: {filename}")
    print(f"     Size: {os.path.getsize(filename) / 1024:.1f} KB")
    print(f"     Sheets: Cover | City Summary | Raw Sensor Data | AI Decisions")
    print("=" * 60)


if __name__ == "__main__":
    main()
